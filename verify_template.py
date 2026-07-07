"""
Sanity-check script for the MiiR Routing Assistant project.

Run with:    python3 verify_template.py

If this prints sheet names, row counts, and a few sample products,
your Python environment is ready for the build session.
"""

import sys
from pathlib import Path

TEMPLATE_NAME = "MiiR_Printer_Availability_Template_v6.xlsx"

# --- Check libraries are installed ---
print("=" * 60)
print("MiiR Routing Assistant — Environment Sanity Check")
print("=" * 60)

try:
    import pandas as pd
    print(f"  pandas: OK  (version {pd.__version__})")
except ImportError:
    print("  pandas: NOT INSTALLED — run: pip install pandas")
    sys.exit(1)

try:
    import openpyxl
    print(f"  openpyxl: OK  (version {openpyxl.__version__})")
except ImportError:
    print("  openpyxl: NOT INSTALLED — run: pip install openpyxl")
    sys.exit(1)

try:
    import streamlit
    print(f"  streamlit: OK  (version {streamlit.__version__})")
except ImportError:
    print("  streamlit: NOT INSTALLED — run: pip install streamlit")
    sys.exit(1)

# pgeocode is optional (zip-code proximity sort)
try:
    import pgeocode
    print(f"  pgeocode: OK  (version {pgeocode.__version__})")
except ImportError:
    print("  pgeocode: NOT INSTALLED — run: pip install pgeocode (zip-proximity sort will be disabled)")

# --- Check the template file is in the right place ---
template_path = Path(__file__).parent / TEMPLATE_NAME
print()
print(f"Looking for template at: {template_path}")
if not template_path.exists():
    print(f"  NOT FOUND. Copy {TEMPLATE_NAME} into this folder.")
    sys.exit(1)
print(f"  Template found: {template_path.stat().st_size:,} bytes")

# --- Read the template ---
print()
print("Reading template sheets:")
print("-" * 60)
wb = openpyxl.load_workbook(template_path, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"  {sheet_name:24}  ({ws.max_row:3} rows × {ws.max_column:2} cols)")

# --- Spot-check a few sheets ---
print()
print("Spot-check — first 5 products:")
print("-" * 60)
products = wb["PRODUCTS"]
shown = 0
for row in products.iter_rows(values_only=True, min_row=5):
    if row[0] and not str(row[0]).startswith("Products"):
        print(f"  {row[0]}")
        shown += 1
        if shown >= 5:
            break

print()
print("Spot-check — decorators:")
print("-" * 60)
decorators = wb["DECORATORS"]
for row in decorators.iter_rows(values_only=True, min_row=5):
    if row[0] and not str(row[0]).startswith(("Source", "Decorator Network")) and row[0] != "Decorator":
        region = row[6] if row[6] else "(no region)"
        zip_code = row[7] if row[7] else "(no zip)"
        print(f"  {str(row[0]):12}  Region: {region:12}  Zip: {zip_code}")

print()
print("Spot-check — first 5 decoration SKUs from SKU_MAP:")
print("-" * 60)
sku_map = wb["SKU_MAP"]
shown = 0
for row in sku_map.iter_rows(values_only=True, min_row=5):
    if row[0] and row[5] == "Decoration":  # col F = Router Category
        deco_type = row[6] if row[6] else "(none)"
        print(f"  {str(row[0]):14}  {str(row[1])[:35]:35}  -> {deco_type}")
        shown += 1
        if shown >= 5:
            break

print()
print("=" * 60)
print("All checks passed. Environment is ready for the build session.")
print("=" * 60)
